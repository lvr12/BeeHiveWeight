import json
from openpyxl import Workbook
from datetime import datetime
import os

def export_to_excel(historique):
    # Déterminer l'année courante
    annee = datetime.now().year
    nom_fichier = f"rendement_{annee}.xlsx"

    # Supprimer l'ancien fichier s'il existe
    if os.path.exists(nom_fichier):
        os.remove(nom_fichier)

    # Créer un nouveau classeur Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Rendement"

    # Titres des colonnes
    ws['A1'] = "Poids"
    ws['B1'] = ""  # colonne vide
    ws['C1'] = "Jour"
    ws['D1'] = ""  # colonne vide
    ws['E1'] = "Heure"

    # Ajouter les données
    for i, entry in enumerate(historique, start=2):
        dt = datetime.fromisoformat(entry["heure"])
        ws[f"A{i}"] = entry["poids"]
        ws[f"B{i}"] = ""
        ws[f"C{i}"] = dt.date()
        ws[f"D{i}"] = ""
        ws[f"E{i}"] = dt.hour  # juste l'heure, pas minutes ni secondes

    # Enregistrer le fichier
    wb.save(nom_fichier)
    return nom_fichier  # ← très important pour le téléchargement
